# Simulador del bot de Cambio Atlántico (casa de cambio USD / EUR).
# Trabajo Práctico Integrador - Organización Empresarial.
#
# El "estado" es un simple texto ("REPOSO", "MENU_MONEDA", etc.) que le da memoria
# al bot, igual que en el diagrama de la maquina de estados.
#
# Requisitos:  pip install openpyxl
# Ejecutar:    python bot_cambio.py

from openpyxl import load_workbook
from datetime import datetime

ARCHIVO = "CambioAtlantico_BD.xlsx"
MAX_INTENTOS = 3
OPERADOR = "+54 9 11 0000-0000"


def leer_cotizaciones():
    libro = load_workbook(ARCHIVO, data_only=True)
    hoja = libro["Cotizaciones"]
    cotizaciones = {}
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        moneda = fila[0]
        if moneda:
            cotizaciones[moneda] = {"compra": float(fila[1]), "venta": float(fila[2])}
    return cotizaciones


def leer_stock():
    libro = load_workbook(ARCHIVO, data_only=True)
    hoja = libro["Stock"]
    stock = {}
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        moneda = fila[0]
        if moneda:
            stock[moneda] = float(fila[1]) if fila[1] is not None else 0
    return stock


def hay_disponibilidad(moneda, operacion, monto):
    stock = leer_stock()
    cotizaciones = leer_cotizaciones()
    if operacion == "comprar":
        return stock[moneda] >= monto
    else:
        pesos_necesarios = monto * cotizaciones[moneda]["compra"]
        return stock["ARS"] >= pesos_necesarios


def calcular_total(moneda, operacion, monto):
    cotizaciones = leer_cotizaciones()
    if operacion == "comprar":
        tasa = cotizaciones[moneda]["venta"]
    else:
        tasa = cotizaciones[moneda]["compra"]
    return float(monto) * float(tasa), float(tasa)


def guardar_operacion(operacion, moneda, monto):
    total, tasa = calcular_total(moneda, operacion, monto)
    libro = load_workbook(ARCHIVO)
    hoja_op = libro["Operaciones"]
    nuevo_id = hoja_op.max_row
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
    hoja_op.append([
        nuevo_id, fecha, "cliente", operacion,
        moneda, int(monto), float(tasa), float(total), "COMPLETADA"
    ])
    hoja_st = libro["Stock"]
    for fila in range(2, hoja_st.max_row + 1):
        nombre = hoja_st.cell(fila, 1).value
        valor_actual = float(hoja_st.cell(fila, 2).value or 0)
        if operacion == "comprar":
            if nombre == moneda:
                hoja_st.cell(fila, 2).value = valor_actual - int(monto)
            if nombre == "ARS":
                hoja_st.cell(fila, 2).value = valor_actual + float(total)
        else:
            if nombre == moneda:
                hoja_st.cell(fila, 2).value = valor_actual + int(monto)
            if nombre == "ARS":
                hoja_st.cell(fila, 2).value = valor_actual - float(total)
    try:
        libro.save(ARCHIVO)
        return total, tasa
    except PermissionError:
        print("\n[ERROR] Cerrá el archivo Excel e intentá de nuevo.\n")
        return None, None


def formatear_precio(valor):
    """Muestra el precio sin decimales si es entero, con decimales si no."""
    if valor == int(valor):
        return str(int(valor))
    return str(round(valor, 2))


# ----------------------------- Programa principal -----------------------------

print("=== Bot Cambio Atlántico ===")
print("Escribí 'operar' para comenzar, o 'salir' para terminar.\n")

estado = "REPOSO"
moneda = ""
operacion = ""
monto = 0
intentos = 0

while True:
    texto = input("Usuario: ").strip()
    cliente = texto.lower()

    if cliente == "salir":
        print("Bot: ¡Hasta luego!")
        break

    if cliente == "operar":
        estado = "MENU_MONEDA"
        moneda = ""
        operacion = ""
        monto = 0
        intentos = 0
        print("Bot: Bienvenido a Cambio Atlántico. ¿Con qué moneda desea operar? (USD o EUR)")
        continue

    if estado == "REPOSO":
        print("Bot: Escribí 'operar' para comenzar.")

    elif estado == "MENU_MONEDA":
        if cliente.upper() in ("USD", "EUR"):
            moneda = cliente.upper()
            intentos = 0
            estado = "MENU_OPERACION"
            cot = leer_cotizaciones()[moneda]
            print("Bot: Cotización " + moneda + ": compra $" + formatear_precio(cot["compra"]) + " / venta $" + formatear_precio(cot["venta"]))
            print("Bot: ¿Querés comprar / vender o solo consultar?")
        else:
            intentos += 1
            if intentos >= MAX_INTENTOS:
                print("Bot: Te derivo con un operador: " + OPERADOR + ". Escribí 'operar' para volver a intentarlo.")
                estado = "REPOSO"
                intentos = 0
            else:
                print("Bot: Por ahora solo operamos USD y EUR. Elegí una de las dos.")

    elif estado == "MENU_OPERACION":
        if cliente in ("comprar", "vender"):
            operacion = cliente
            intentos = 0
            estado = "ESPERA_MONTO"
            print("Bot: ¿Qué monto de " + moneda + " querés " + operacion + "?")
        elif cliente in ("consultar", "no", "salir"):
            estado = "REPOSO"
            intentos = 0
            print("Bot: ¡Listo! Esa es la cotización de hoy. Escribí 'operar' cuando quieras comenzar.")
        else:
            intentos += 1
            if intentos >= MAX_INTENTOS:
                print("Bot: Te derivo con un operador: " + OPERADOR + ". Escribí 'operar' para volver a intentarlo.")
                estado = "REPOSO"
                intentos = 0
            else:
                print("Bot: Elegí: comprar, vender o consultar.")

    elif estado == "ESPERA_MONTO":
        if not cliente.isdigit():
            intentos += 1
            if intentos >= MAX_INTENTOS:
                print("Bot: Te derivo con un operador: " + OPERADOR + ". Escribí 'operar' para volver a intentarlo.")
                estado = "REPOSO"
                intentos = 0
            else:
                print("Bot: Ingresá el monto en números, por ejemplo: 100.")
        elif int(cliente) <= 0:
            print("Bot: El monto debe ser mayor a cero.")
        elif not hay_disponibilidad(moneda, operacion, int(cliente)):
            estado = "MENU_OPERACION"
            intentos = 0
            print("Bot: No tenemos esa cantidad disponible. Elegí otra operación o usá 'operar'.")
        else:
            monto = int(cliente)
            intentos = 0
            total, tasa = calcular_total(moneda, operacion, monto)
            estado = "ESPERA_CONFIRMACION"
            print(
                "Bot: " + operacion.capitalize() + " de " + str(monto) + " " + moneda +
                " a $" + formatear_precio(tasa) + ". Total: $" + formatear_precio(total) + ". ¿Confirmás? (sí/no)"
            )

    elif estado == "ESPERA_CONFIRMACION":
        if cliente in ("si", "sí", "s"):
            total, tasa = guardar_operacion(operacion, moneda, monto)
            estado = "REPOSO"  # vuelve a REPOSO siempre, haya guardado o no
            intentos = 0
            if total is not None:
                print("Bot: Operación registrada. Total: $" + formatear_precio(total) + ".")
                print("Bot: Gracias por operar con Cambio Atlántico. Escribí 'operar' para otra operación.")
        elif cliente in ("no", "n", "cancelar"):
            estado = "REPOSO"
            intentos = 0
            print("Bot: Operación cancelada. Escribí 'operar' cuando quieras.")
        else:
            print("Bot: Respondé sí o no para confirmar.")
